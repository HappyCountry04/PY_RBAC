from typing import Annotated, Any

import io
import os
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_current_user, require_perm
from app.core.response import success, table
from app.core.security import get_password_hash, verify_password
from app.db.session import get_db
from app.models import (
    SysDept,
    SysPost,
    SysRole,
    SysUser,
    SysUserPost,
    SysUserRole,
)
from app.services.rbac import LoginUser
from app.api.routes.system.common import (
    _parse_import_rows,
    apply_data_scope,
    apply_time_range,
    build_dept_tree,
    collect_child_dept_ids,
    current_name,
    dept_tree,
    ensure_exists,
    ensure_unique,
    field,
    has_field,
    page_params,
    page_query,
    replace_rows,
    serialize_post,
    serialize_role,
    serialize_user,
    sync_user_roles_posts,
    xlsx_export,
)

router = APIRouter()


@router.get("/user/list")
async def user_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    user_name: str | None = Query(None, alias="userName"),
    phonenumber: str | None = None,
    status: str | None = None,
    dept_id: int | None = Query(None, alias="deptId"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysUser).options(selectinload(SysUser.dept)).where(SysUser.del_flag == "0").order_by(SysUser.user_id)
    stmt = await apply_data_scope(db, stmt, login_user, SysUser.dept_id, SysUser.user_id)
    if user_name:
        stmt = stmt.where(SysUser.user_name.ilike(f"%{user_name}%"))
    if phonenumber:
        stmt = stmt.where(SysUser.phonenumber.ilike(f"%{phonenumber}%"))
    if status:
        stmt = stmt.where(SysUser.status == status)
    if dept_id:
        child_dept_ids = await collect_child_dept_ids(db, dept_id)
        stmt = stmt.where(SysUser.dept_id.in_([dept_id, *child_dept_ids]))
    stmt = apply_time_range(stmt, SysUser.create_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_user(row) for row in rows], total)


@router.get("/user/deptTree")
async def user_dept_tree(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:list"))],
):
    depts = list(
        (await db.execute(select(SysDept).where(SysDept.del_flag == "0").order_by(SysDept.parent_id, SysDept.order_num))).scalars()
    )
    return success([dept_tree(node) for node in build_dept_tree(depts, 0)])


@router.get("/user/")
async def user_add_options(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    roles = list((await db.execute(select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort))).scalars())
    posts = list((await db.execute(select(SysPost).order_by(SysPost.post_sort))).scalars())
    return success(roles=[serialize_role(role) for role in roles], posts=[serialize_post(post) for post in posts])


@router.get("/user/importTemplate")
async def user_import_template(
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    headers = ["userName", "nickName", "deptId", "email", "phonenumber", "sex", "status", "password"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    example = ["zhangsan", "张三", "100", "zhangsan@example.com", "13800000000", "0", "0", "123456"]
    for col_idx, value in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=value)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"},
    )


@router.post("/user/importData")
async def user_import_data(
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:add"))],
    file: UploadFile = File(...),
    update_support: bool = Query(False, alias="updateSupport"),
):
    raw = await file.read()
    filename = file.filename or ""
    rows_data = _parse_import_rows(raw, filename)
    if not rows_data or "userName" not in rows_data[0]:
        raise HTTPException(status_code=400, detail="导入文件必须包含 userName 表头")

    created = 0
    updated = 0
    skipped: list[str] = []
    now = datetime.now()
    for line_no, row in enumerate(rows_data, start=2):
        user_name = (row.get("userName") or "").strip()
        if not user_name:
            skipped.append(f"第 {line_no} 行缺少 userName")
            continue
        user = await db.scalar(select(SysUser).where(SysUser.user_name == user_name))
        if user and not update_support:
            skipped.append(f"{user_name} 已存在")
            continue
        if not user:
            user = SysUser(
                user_name=user_name,
                password=get_password_hash((row.get("password") or "123456").strip() or "123456"),
                create_by=current_name(login_user),
                create_time=now,
                del_flag="0",
            )
            db.add(user)
            created += 1
        else:
            updated += 1
        if row.get("deptId"):
            user.dept_id = int(row["deptId"])
        user.nick_name = (row.get("nickName") or user_name).strip()
        user.email = (row.get("email") or "").strip()
        user.phonenumber = (row.get("phonenumber") or "").strip()
        user.sex = (row.get("sex") or "0").strip()
        user.status = (row.get("status") or "0").strip()
        user.remark = row.get("remark")
        user.update_by = current_name(login_user)
        user.update_time = now
    await db.commit()
    return success()


# UPLOAD_DIR for avatars
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "uploads", "avatars")
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.get("/user/profile")
async def user_profile(
    login_user: Annotated[LoginUser, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user = login_user.user
    posts = list(
        (await db.execute(
            select(SysPost).join(SysUserPost, SysPost.post_id == SysUserPost.post_id)
            .where(SysUserPost.user_id == user.user_id)
        )).scalars()
    )
    roles = list(
        (await db.execute(
            select(SysRole).join(SysUserRole, SysRole.role_id == SysUserRole.role_id)
            .where(SysUserRole.user_id == user.user_id)
        )).scalars()
    )
    return success(
        data=serialize_user(user),
        roleGroup="、".join(r.role_name for r in roles) or "无角色",
        postGroup="、".join(p.post_name for p in posts) or "无岗位",
        roles=[serialize_role(r) for r in roles],
        postIds=[p.post_id for p in posts],
        roleIds=[r.role_id for r in roles],
    )


@router.put("/user/profile")
async def user_profile_update(
    body: dict[str, Any],
    login_user: Annotated[LoginUser, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user = login_user.user
    for attr, key in [
        ("nick_name", "nickName"),
        ("phonenumber", "phonenumber"),
        ("email", "email"),
        ("sex", "sex"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(user, attr, value)
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    await db.commit()
    return success()


@router.put("/user/profile/updatePwd")
async def user_profile_update_pwd(
    body: dict[str, Any],
    login_user: Annotated[LoginUser, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    old_password = str(field(body, "oldPassword", ""))
    new_password = str(field(body, "newPassword", ""))
    if not verify_password(old_password, login_user.user.password):
        raise HTTPException(status_code=400, detail="旧密码不正确")
    login_user.user.password = get_password_hash(new_password)
    login_user.user.update_by = current_name(login_user)
    login_user.user.update_time = datetime.now()
    await db.commit()
    return success()


@router.post("/user/profile/avatar")
async def user_profile_avatar(
    avatarfile: UploadFile = File(...),
    login_user: LoginUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ext = os.path.splitext(avatarfile.filename or "avatar.png")[1] or ".png"
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    content = await avatarfile.read()
    with open(filepath, "wb") as f:
        f.write(content)
    login_user.user.avatar = f"/uploads/avatars/{filename}"
    login_user.user.update_by = current_name(login_user)
    login_user.user.update_time = datetime.now()
    await db.commit()
    return success(imgUrl=f"/uploads/avatars/{filename}")


@router.get("/user/authRole/{user_id}")
async def user_auth_role_detail(
    user_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    user = await db.get(SysUser, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="数据不存在")
    assigned_ids = set(
        (await db.execute(select(SysUserRole.role_id).where(SysUserRole.user_id == user_id))).scalars()
    )
    all_roles = list((await db.execute(select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort))).scalars())
    return success(
        user=serialize_user(user),
        roles=[{**serialize_role(r), "flag": r.role_id in assigned_ids} for r in all_roles],
    )


@router.get("/user/export")
async def user_export(
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
    user_name: str | None = Query(None, alias="userName"),
    phonenumber: str | None = None,
    status: str | None = None,
    dept_id: int | None = Query(None, alias="deptId"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysUser).options(selectinload(SysUser.dept)).where(SysUser.del_flag == "0").order_by(SysUser.user_id)
    stmt = await apply_data_scope(db, stmt, login_user, SysUser.dept_id, SysUser.user_id)
    if user_name:
        stmt = stmt.where(SysUser.user_name.ilike(f"%{user_name}%"))
    if phonenumber:
        stmt = stmt.where(SysUser.phonenumber.ilike(f"%{phonenumber}%"))
    if status:
        stmt = stmt.where(SysUser.status == status)
    if dept_id:
        child_dept_ids = await collect_child_dept_ids(db, dept_id)
        stmt = stmt.where(SysUser.dept_id.in_([dept_id, *child_dept_ids]))
    stmt = apply_time_range(stmt, SysUser.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars().unique())
    data = [
        [
            str(u.user_id), u.user_name, u.nick_name,
            u.dept.dept_name if u.dept else "",
            u.phonenumber, u.email,
            "男" if u.sex == "0" else "女" if u.sex == "1" else "未知",
            "正常" if u.status == "0" else "停用",
            str(u.create_time or ""), u.remark or ""
        ]
        for u in rows
    ]
    return await xlsx_export(
        "user_export.xlsx",
        ["用户ID", "账号", "昵称", "部门", "手机", "邮箱", "性别", "状态", "创建时间", "备注"],
        data,
    )


@router.get("/user/{user_id}")
async def user_detail(
    user_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    user = (
        await db.execute(
            select(SysUser)
            .options(selectinload(SysUser.roles), selectinload(SysUser.posts))
            .where(SysUser.user_id == user_id)
        )
    ).scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="数据不存在")
    role_names = "、".join(role.role_name for role in user.roles) or "无角色"
    post_names = "、".join(post.post_name for post in user.posts) or "无岗位"
    role_ids = [role.role_id for role in user.roles]
    post_ids = [post.post_id for post in user.posts]
    roles = list((await db.execute(select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort))).scalars())
    posts = list((await db.execute(select(SysPost).order_by(SysPost.post_sort))).scalars())
    data = serialize_user(user)
    data["roleNames"] = role_names
    data["postNames"] = post_names
    return success(
        data=data,
        roleIds=role_ids,
        postIds=post_ids,
        roles=[serialize_role(role) for role in roles],
        posts=[serialize_post(post) for post in posts],
    )


@router.post("/user")
async def user_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:add"))],
):
    now = datetime.now()
    password = str(field(body, "password", "123456"))
    await ensure_unique(db, SysUser, SysUser.user_name, field(body, "userName"), "用户账号已存在")
    await ensure_unique(db, SysUser, SysUser.phonenumber, field(body, "phonenumber"), "手机号已存在")
    await ensure_unique(db, SysUser, SysUser.email, field(body, "email"), "邮箱已存在")
    user = SysUser(
        dept_id=field(body, "deptId"),
        user_name=field(body, "userName"),
        nick_name=field(body, "nickName", field(body, "userName")),
        email=field(body, "email", ""),
        phonenumber=field(body, "phonenumber", ""),
        sex=field(body, "sex", "0"),
        password=get_password_hash(password),
        status=field(body, "status", "0"),
        remark=field(body, "remark"),
        create_by=current_name(login_user),
        create_time=now,
    )
    db.add(user)
    await db.flush()
    if has_field(body, "roleIds") or has_field(body, "postIds"):
        await sync_user_roles_posts(
            db,
            user.user_id,
            field(body, "roleIds", []),
            field(body, "postIds", []),
        )
    await db.commit()
    return success()


@router.put("/user")
async def user_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:edit"))],
):
    user_id = int(field(body, "userId"))
    user = await ensure_exists(db, SysUser, user_id)
    if user.is_admin and not login_user.user.is_admin:
        raise HTTPException(status_code=403, detail="不能修改超级管理员")
    await ensure_unique(db, SysUser, SysUser.user_name, field(body, "userName"), "用户账号已存在", SysUser.user_id, user_id)
    await ensure_unique(db, SysUser, SysUser.phonenumber, field(body, "phonenumber"), "手机号已存在", SysUser.user_id, user_id)
    await ensure_unique(db, SysUser, SysUser.email, field(body, "email"), "邮箱已存在", SysUser.user_id, user_id)
    for attr, key in [
        ("dept_id", "deptId"),
        ("user_name", "userName"),
        ("nick_name", "nickName"),
        ("email", "email"),
        ("phonenumber", "phonenumber"),
        ("sex", "sex"),
        ("status", "status"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(user, attr, value)
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    if has_field(body, "roleIds") or has_field(body, "postIds"):
        await sync_user_roles_posts(
            db,
            user.user_id,
            field(body, "roleIds", []),
            field(body, "postIds", []),
        )
    await db.commit()
    return success()


@router.put("/user/resetPwd")
async def user_reset_pwd(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:resetPwd"))],
):
    user = await ensure_exists(db, SysUser, int(field(body, "userId")))
    user.password = get_password_hash(str(field(body, "password")))
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    await db.commit()
    return success()


@router.put("/user/changeStatus")
async def user_change_status(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:edit"))],
):
    user = await ensure_exists(db, SysUser, int(field(body, "userId")))
    if user.is_admin:
        raise HTTPException(status_code=400, detail="不能停用超级管理员")
    user.status = str(field(body, "status", user.status))
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/user/{user_ids}")
async def user_remove(
    user_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:remove"))],
):
    ids = [int(item) for item in user_ids.split(",") if item]
    if 1 in ids:
        raise HTTPException(status_code=400, detail="不能删除超级管理员")
    rows = (await db.execute(select(SysUser).where(SysUser.user_id.in_(ids)))).scalars()
    for user in rows:
        user.del_flag = "2"
    await db.commit()
    return success()


@router.put("/user/authRole")
async def user_auth_role(
    user_id: int | None = Query(None, alias="userId"),
    user_id_snake: int | None = Query(None, alias="user_id"),
    role_ids: str = Query("", alias="roleIds"),
    db: AsyncSession = Depends(get_db),
    _login_user: LoginUser = Depends(require_perm("system:user:edit")),
):
    user_id = user_id or user_id_snake
    if user_id is None:
        raise HTTPException(status_code=400, detail="缺少用户ID")
    ids = [int(item) for item in role_ids.split(",") if item]
    await replace_rows(
        db,
        SysUserRole,
        SysUserRole.user_id == user_id,
        [SysUserRole(user_id=user_id, role_id=role_id) for role_id in ids],
    )
    await db.commit()
    return success()
