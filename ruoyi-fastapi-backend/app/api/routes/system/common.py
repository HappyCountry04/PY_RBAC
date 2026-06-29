from typing import Any, TypeVar

from datetime import datetime
import io

from fastapi import HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import Select, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import (
    SysConfig,
    SysDept,
    SysDictData,
    SysDictType,
    SysLogininfor,
    SysMenu,
    SysOperLog,
    SysPost,
    SysRole,
    SysUser,
    SysUserPost,
    SysUserRole,
)
from app.services.cache import dict_cache_set
from app.services.rbac import LoginUser, apply_data_scope

T = TypeVar("T")


async def _refresh_dict_cache(db: AsyncSession, dict_type: str) -> None:
    """刷新单个字典类型的 Redis 缓存"""
    rows = list(
        (
            await db.execute(
                select(SysDictData)
                .where(SysDictData.dict_type == dict_type, SysDictData.status == "0")
                .order_by(SysDictData.dict_sort)
            )
        ).scalars()
    )
    await dict_cache_set(dict_type, [serialize_dict_data(row) for row in rows])


async def page_query(
    db: AsyncSession,
    stmt: Select,
    page_num: int,
    page_size: int,
) -> tuple[list[Any], int]:
    total_stmt = select(func.count()).select_from(stmt.order_by(None).subquery())
    total = (await db.execute(total_stmt)).scalar_one()
    result = await db.execute(stmt.offset((page_num - 1) * page_size).limit(page_size))
    return list(result.scalars().unique()), total


def page_params(
    page_num: int = Query(1, alias="pageNum", ge=1),
    page_size: int = Query(10, alias="pageSize", ge=1, le=200),
) -> tuple[int, int]:
    return page_num, page_size


def field(data: dict[str, Any], name: str, default: Any = None) -> Any:
    snake = []
    for char in name:
        if char.isupper():
            snake.append("_")
            snake.append(char.lower())
        else:
            snake.append(char)
    return data.get(name, data.get("".join(snake), default))


def has_field(data: dict[str, Any], name: str) -> bool:
    snake = []
    for char in name:
        if char.isupper():
            snake.append("_")
            snake.append(char.lower())
        else:
            snake.append(char)
    return name in data or "".join(snake) in data


def current_name(login_user: LoginUser) -> str:
    return login_user.user.user_name


def apply_time_range(stmt: Select, column, begin_time: str | None, end_time: str | None) -> Select:
    if begin_time:
        stmt = stmt.where(column >= datetime.fromisoformat(begin_time))
    if end_time:
        end_value = datetime.fromisoformat(end_time)
        if len(end_time) <= 10:
            end_value = end_value.replace(hour=23, minute=59, second=59)
        stmt = stmt.where(column <= end_value)
    return stmt


async def ensure_exists(db: AsyncSession, model, pk: Any):
    item = await db.get(model, pk)
    if item is None:
        raise HTTPException(status_code=404, detail="数据不存在")
    return item


async def ensure_unique(db: AsyncSession, model, column, value: Any, message: str, pk_column=None, pk_value: Any = None) -> None:
    if value in (None, ""):
        return
    stmt = select(func.count()).select_from(model).where(column == value)
    if pk_column is not None and pk_value not in (None, ""):
        stmt = stmt.where(pk_column != pk_value)
    if await db.scalar(stmt):
        raise HTTPException(status_code=400, detail=message)


async def replace_rows(db: AsyncSession, model, where_clause, rows: list[Any]) -> None:
    await db.execute(delete(model).where(where_clause))
    db.add_all(rows)


def _parse_import_rows(raw: bytes, filename: str) -> list[dict[str, str]]:
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None)
        if not headers_row:
            wb.close()
            return []
        headers = [str(h).strip() if h else "" for h in headers_row]
        result: list[dict[str, str]] = []
        for row in rows_iter:
            values = [str(v).strip() if v is not None else "" for v in row]
            if not any(values):
                continue
            result.append(dict(zip(headers, values, strict=False)))
        wb.close()
        return result
    elif filename.lower().endswith(".csv"):
        import csv
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader]
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 文件")


async def xlsx_export(filename: str, headers: list[str], rows: list[list[str]]) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_data in rows:
            cell_value = str(row_data[col_idx - 1]) if col_idx - 1 < len(row_data) else ""
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def model_dict(obj) -> dict:
    return {col.name: getattr(obj, col.name) for col in obj.__table__.columns}


def serialize_user(user: SysUser) -> dict:
    return {
        "userId": user.user_id,
        "deptId": user.dept_id,
        "deptName": user.dept.dept_name if user.dept else "",
        "dept": {"deptName": user.dept.dept_name} if user.dept else None,
        "userName": user.user_name,
        "nickName": user.nick_name,
        "email": user.email,
        "phonenumber": user.phonenumber,
        "sex": user.sex,
        "avatar": user.avatar,
        "status": user.status,
        "remark": user.remark,
        "createTime": user.create_time,
    }


def serialize_role(role: SysRole) -> dict:
    data = model_dict(role)
    data.update({
        "roleId": role.role_id,
        "roleName": role.role_name,
        "roleKey": role.role_key,
        "dataScope": role.data_scope,
        "roleSort": role.role_sort,
        "createTime": role.create_time,
        "updateTime": role.update_time,
    })
    return data


def serialize_menu(menu: SysMenu) -> dict:
    return {
        "menuId": menu.menu_id,
        "menuName": menu.menu_name,
        "parentId": menu.parent_id,
        "orderNum": menu.order_num,
        "path": menu.path,
        "component": menu.component,
        "query": menu.query,
        "routeName": menu.route_name,
        "isFrame": menu.is_frame,
        "isCache": menu.is_cache,
        "menuType": menu.menu_type,
        "visible": menu.visible,
        "status": menu.status,
        "perms": menu.perms,
        "icon": menu.icon,
        "remark": menu.remark,
        "createTime": menu.create_time,
        "updateTime": menu.update_time,
    }


def serialize_dept(dept: SysDept) -> dict:
    return {
        "deptId": dept.dept_id,
        "parentId": dept.parent_id,
        "ancestors": dept.ancestors,
        "deptName": dept.dept_name,
        "orderNum": dept.order_num,
        "leader": dept.leader,
        "phone": dept.phone,
        "email": dept.email,
        "status": dept.status,
        "createTime": dept.create_time,
        "updateTime": dept.update_time,
    }


def serialize_post(post: SysPost) -> dict:
    data = model_dict(post)
    data.update({
        "postId": post.post_id,
        "postCode": post.post_code,
        "postName": post.post_name,
        "postSort": post.post_sort,
        "createTime": post.create_time,
        "updateTime": post.update_time,
    })
    return data


def serialize_config(config: SysConfig) -> dict:
    data = model_dict(config)
    data.update({
        "configId": config.config_id,
        "configName": config.config_name,
        "configKey": config.config_key,
        "configValue": config.config_value,
        "configType": config.config_type,
        "createTime": config.create_time,
        "updateTime": config.update_time,
    })
    return data


def serialize_dict_type(item: SysDictType) -> dict:
    data = model_dict(item)
    data.update({
        "dictId": item.dict_id,
        "dictName": item.dict_name,
        "dictType": item.dict_type,
        "createTime": item.create_time,
        "updateTime": item.update_time,
    })
    return data


def serialize_dict_data(item: SysDictData) -> dict:
    data = model_dict(item)
    data.update({
        "dictCode": item.dict_code,
        "dictSort": item.dict_sort,
        "dictLabel": item.dict_label,
        "dictValue": item.dict_value,
        "dictType": item.dict_type,
        "cssClass": item.css_class,
        "listClass": item.list_class,
        "isDefault": item.is_default,
        "createTime": item.create_time,
        "updateTime": item.update_time,
    })
    return data


def tree_select(menu: SysMenu) -> dict:
    return {
        "id": menu.menu_id,
        "label": menu.menu_name,
        "children": [tree_select(child) for child in getattr(menu, "children", [])],
    }


def data_scope_label_str(value: str) -> str:
    return {
        "1": "全部数据权限",
        "2": "自定义数据权限",
        "3": "本部门数据权限",
        "4": "本部门及以下数据权限",
        "5": "仅本人数据权限",
    }.get(value, value)


def business_type_label_str(value: int) -> str:
    return {
        0: "其他",
        1: "新增",
        2: "修改",
        3: "删除",
        4: "授权",
        5: "导出",
        6: "导入",
        7: "强退",
        8: "清空",
        9: "生成代码",
    }.get(value, str(value))


async def user_export(
    db: AsyncSession,
    login_user: LoginUser,
    user_name: str | None = Query(None, alias="userName"),
    phonenumber: str | None = None,
    status: str | None = None,
    dept_id: int | None = Query(None, alias="deptId"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysUser).options(selectinload(SysUser.dept)).where(SysUser.del_flag == "0").order_by(SysUser.user_id)
    stmt = await apply_data_scope(db, stmt, login_user, SysUser.dept_id, SysUser.user_id)
    if user_name:
        stmt = stmt.where(SysUser.user_name.ilike(f"%{user_name}%"))
    if phonenumber:
        stmt = stmt.where(SysUser.phonenumber.ilike(f"%{phonenumber}%"))
    if status:
        stmt = stmt.where(SysUser.status == status)
    if dept_id:
        child_dept_ids = await collect_child_dept_ids(db, dept_id)
        stmt = stmt.where(SysUser.dept_id.in_([dept_id, *child_dept_ids]))
    stmt = apply_time_range(stmt, SysUser.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars().unique())
    data = [
        [
            str(u.user_id), u.user_name, u.nick_name,
            u.dept.dept_name if u.dept else "",
            u.phonenumber, u.email,
            "男" if u.sex == "0" else "女" if u.sex == "1" else "未知",
            "正常" if u.status == "0" else "停用",
            str(u.create_time or ""), u.remark or ""
        ]
        for u in rows
    ]
    return await xlsx_export(
        "user_export.xlsx",
        ["用户ID", "账号", "昵称", "部门", "手机", "邮箱", "性别", "状态", "创建时间", "备注"],
        data,
    )


async def role_export(
    db: AsyncSession,
    role_name: str | None = Query(None, alias="roleName"),
    role_key: str | None = Query(None, alias="roleKey"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort)
    if role_name:
        stmt = stmt.where(SysRole.role_name.ilike(f"%{role_name}%"))
    if role_key:
        stmt = stmt.where(SysRole.role_key.ilike(f"%{role_key}%"))
    if status:
        stmt = stmt.where(SysRole.status == status)
    stmt = apply_time_range(stmt, SysRole.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [
            str(r.role_id), r.role_name, r.role_key, str(r.role_sort),
            data_scope_label_str(r.data_scope),
            "正常" if r.status == "0" else "停用",
            str(r.create_time or ""), r.remark or ""
        ]
        for r in rows
    ]
    return await xlsx_export(
        "role_export.xlsx",
        ["角色ID", "角色名称", "权限字符", "排序", "数据范围", "状态", "创建时间", "备注"],
        data,
    )


async def post_export(
    db: AsyncSession,
    post_code: str | None = Query(None, alias="postCode"),
    post_name: str | None = Query(None, alias="postName"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysPost).order_by(SysPost.post_sort)
    if post_code:
        stmt = stmt.where(SysPost.post_code.ilike(f"%{post_code}%"))
    if post_name:
        stmt = stmt.where(SysPost.post_name.ilike(f"%{post_name}%"))
    if status:
        stmt = stmt.where(SysPost.status == status)
    stmt = apply_time_range(stmt, SysPost.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(p.post_id), p.post_code, p.post_name, str(p.post_sort),
         "正常" if p.status == "0" else "停用",
         str(p.create_time or ""), p.remark or ""]
        for p in rows
    ]
    return await xlsx_export(
        "post_export.xlsx",
        ["岗位ID", "岗位编码", "岗位名称", "排序", "状态", "创建时间", "备注"],
        data,
    )


async def config_export(
    db: AsyncSession,
    config_name: str | None = Query(None, alias="configName"),
    config_key: str | None = Query(None, alias="configKey"),
    config_type: str | None = Query(None, alias="configType"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysConfig).order_by(SysConfig.config_id)
    if config_name:
        stmt = stmt.where(SysConfig.config_name.ilike(f"%{config_name}%"))
    if config_key:
        stmt = stmt.where(SysConfig.config_key.ilike(f"%{config_key}%"))
    if config_type:
        stmt = stmt.where(SysConfig.config_type == config_type)
    stmt = apply_time_range(stmt, SysConfig.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(c.config_id), c.config_name or "", c.config_key,
         c.config_value or "", "是" if c.config_type == "Y" else "否",
         str(c.create_time or ""), c.remark or ""]
        for c in rows
    ]
    return await xlsx_export(
        "config_export.xlsx",
        ["参数ID", "参数名称", "参数键名", "参数键值", "系统内置", "创建时间", "备注"],
        data,
    )


async def operlog_export(
    db: AsyncSession,
    title: str | None = None,
    oper_name: str | None = Query(None, alias="operName"),
    business_type: int | None = Query(None, alias="businessType"),
    status: int | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysOperLog).order_by(SysOperLog.oper_id.desc())
    if title:
        stmt = stmt.where(SysOperLog.title.ilike(f"%{title}%"))
    if oper_name:
        stmt = stmt.where(SysOperLog.oper_name.ilike(f"%{oper_name}%"))
    if business_type is not None:
        stmt = stmt.where(SysOperLog.business_type == business_type)
    if status is not None:
        stmt = stmt.where(SysOperLog.status == status)
    stmt = apply_time_range(stmt, SysOperLog.oper_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(o.oper_id), o.title or "", business_type_label_str(o.business_type or 0),
         o.request_method or "", o.oper_name or "", o.oper_url or "",
         o.oper_ip or "", str(o.oper_time or ""), str(o.cost_time or 0),
         "成功" if o.status == 0 else "失败", o.error_msg or ""]
        for o in rows
    ]
    return await xlsx_export(
        "operlog_export.xlsx",
        ["日志编号", "操作模块", "业务类型", "请求方式", "操作人员", "请求URL", "IP地址", "操作时间", "耗时(ms)", "状态", "错误信息"],
        data,
    )


async def logininfor_export(
    db: AsyncSession,
    user_name: str | None = Query(None, alias="userName"),
    ipaddr: str | None = None,
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysLogininfor).order_by(SysLogininfor.info_id.desc())
    if user_name:
        stmt = stmt.where(SysLogininfor.user_name.ilike(f"%{user_name}%"))
    if ipaddr:
        stmt = stmt.where(SysLogininfor.ipaddr.ilike(f"%{ipaddr}%"))
    if status:
        stmt = stmt.where(SysLogininfor.status == status)
    stmt = apply_time_range(stmt, SysLogininfor.login_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(l.info_id), l.user_name or "", l.ipaddr or "", l.login_location or "",
         l.browser or "", l.os or "", "成功" if l.status == "0" else "失败",
         l.msg or "", str(l.login_time or "")]
        for l in rows
    ]
    return await xlsx_export(
        "logininfor_export.xlsx",
        ["访问编号", "用户账号", "IP地址", "登录地点", "浏览器", "操作系统", "状态", "提示消息", "登录时间"],
        data,
    )


async def collect_child_dept_ids(db: AsyncSession, dept_id: int) -> list[int]:
    rows = list((await db.execute(select(SysDept).where(SysDept.del_flag == "0"))).scalars())
    result: list[int] = []

    def collect(parent_id: int) -> None:
        for dept in rows:
            if dept.parent_id == parent_id:
                result.append(dept.dept_id)
                collect(dept.dept_id)

    collect(dept_id)
    return result


def build_dept_tree(depts: list[SysDept], parent_id: int = 0) -> list[SysDept]:
    nodes = [dept for dept in depts if dept.parent_id == parent_id]
    for node in nodes:
        setattr(node, "children", build_dept_tree(depts, node.dept_id))
    return nodes


def dept_tree(dept: SysDept) -> dict:
    return {
        "id": dept.dept_id,
        "label": dept.dept_name,
        "children": [dept_tree(child) for child in getattr(dept, "children", [])],
    }


async def sync_user_roles_posts(
    db: AsyncSession,
    user_id: int,
    role_ids: list[int] | None,
    post_ids: list[int] | None,
) -> None:
    await replace_rows(
        db,
        SysUserRole,
        SysUserRole.user_id == user_id,
        [SysUserRole(user_id=user_id, role_id=int(role_id)) for role_id in role_ids or []],
    )
    await replace_rows(
        db,
        SysUserPost,
        SysUserPost.user_id == user_id,
        [SysUserPost(user_id=user_id, post_id=int(post_id)) for post_id in post_ids or []],
    )
