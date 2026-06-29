from typing import Annotated, Any, TypeVar

from datetime import datetime
import io
import os
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import Select, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_current_user, require_perm
from app.core.response import success, table
from app.core.security import get_password_hash, verify_password
from app.db.session import get_db
from app.models import (
    SysConfig,
    SysDept,
    SysDictData,
    SysDictType,
    SysLogininfor,
    SysMenu,
    SysOperLog,
    SysPost,
    SysRole,
    SysRoleDept,
    SysRoleMenu,
    SysUser,
    SysUserPost,
    SysUserRole,
)
from app.services.rbac import LoginUser, apply_data_scope, build_tree, invalidate_tokens_for_role

router = APIRouter(prefix="/system", tags=["system"])

T = TypeVar("T")


async def page_query(
    db: AsyncSession,
    stmt: Select,
    page_num: int,
    page_size: int,
) -> tuple[list[Any], int]:
    total_stmt = select(func.count()).select_from(stmt.order_by(None).subquery())
    total = (await db.execute(total_stmt)).scalar_one()
    result = await db.execute(stmt.offset((page_num - 1) * page_size).limit(page_size))
    return list(result.scalars().unique()), total


def page_params(
    page_num: int = Query(1, alias="pageNum", ge=1),
    page_size: int = Query(10, alias="pageSize", ge=1, le=200),
) -> tuple[int, int]:
    return page_num, page_size


def field(data: dict[str, Any], name: str, default: Any = None) -> Any:
    snake = []
    for char in name:
        if char.isupper():
            snake.append("_")
            snake.append(char.lower())
        else:
            snake.append(char)
    return data.get(name, data.get("".join(snake), default))


def has_field(data: dict[str, Any], name: str) -> bool:
    snake = []
    for char in name:
        if char.isupper():
            snake.append("_")
            snake.append(char.lower())
        else:
            snake.append(char)
    return name in data or "".join(snake) in data


def current_name(login_user: LoginUser) -> str:
    return login_user.user.user_name


def apply_time_range(stmt: Select, column, begin_time: str | None, end_time: str | None) -> Select:
    if begin_time:
        stmt = stmt.where(column >= datetime.fromisoformat(begin_time))
    if end_time:
        end_value = datetime.fromisoformat(end_time)
        if len(end_time) <= 10:
            end_value = end_value.replace(hour=23, minute=59, second=59)
        stmt = stmt.where(column <= end_value)
    return stmt


async def ensure_exists(db: AsyncSession, model, pk: Any):
    item = await db.get(model, pk)
    if item is None:
        raise HTTPException(status_code=404, detail="数据不存在")
    return item


async def ensure_unique(db: AsyncSession, model, column, value: Any, message: str, pk_column=None, pk_value: Any = None) -> None:
    if value in (None, ""):
        return
    stmt = select(func.count()).select_from(model).where(column == value)
    if pk_column is not None and pk_value not in (None, ""):
        stmt = stmt.where(pk_column != pk_value)
    if await db.scalar(stmt):
        raise HTTPException(status_code=400, detail=message)


async def replace_rows(db: AsyncSession, model, where_clause, rows: list[Any]) -> None:
    await db.execute(delete(model).where(where_clause))
    db.add_all(rows)


@router.get("/user/list")
async def user_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    user_name: str | None = Query(None, alias="userName"),
    phonenumber: str | None = None,
    status: str | None = None,
    dept_id: int | None = Query(None, alias="deptId"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysUser).options(selectinload(SysUser.dept)).where(SysUser.del_flag == "0").order_by(SysUser.user_id)
    stmt = await apply_data_scope(db, stmt, login_user, SysUser.dept_id, SysUser.user_id)
    if user_name:
        stmt = stmt.where(SysUser.user_name.ilike(f"%{user_name}%"))
    if phonenumber:
        stmt = stmt.where(SysUser.phonenumber.ilike(f"%{phonenumber}%"))
    if status:
        stmt = stmt.where(SysUser.status == status)
    if dept_id:
        child_dept_ids = await collect_child_dept_ids(db, dept_id)
        stmt = stmt.where(SysUser.dept_id.in_([dept_id, *child_dept_ids]))
    stmt = apply_time_range(stmt, SysUser.create_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_user(row) for row in rows], total)


@router.get("/user/deptTree")
async def user_dept_tree(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:list"))],
):
    depts = list(
        (await db.execute(select(SysDept).where(SysDept.del_flag == "0").order_by(SysDept.parent_id, SysDept.order_num))).scalars()
    )
    return success([dept_tree(node) for node in build_dept_tree(depts, 0)])


@router.get("/user/")
async def user_add_options(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    roles = list((await db.execute(select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort))).scalars())
    posts = list((await db.execute(select(SysPost).order_by(SysPost.post_sort))).scalars())
    return success(roles=[serialize_role(role) for role in roles], posts=[serialize_post(post) for post in posts])


@router.get("/user/importTemplate")
async def user_import_template(
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    headers = ["userName", "nickName", "deptId", "email", "phonenumber", "sex", "status", "password"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    example = ["zhangsan", "张三", "100", "zhangsan@example.com", "13800000000", "0", "0", "123456"]
    for col_idx, value in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=value)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"},
    )


def _parse_import_rows(raw: bytes, filename: str) -> list[dict[str, str]]:
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None)
        if not headers_row:
            wb.close()
            return []
        headers = [str(h).strip() if h else "" for h in headers_row]
        result: list[dict[str, str]] = []
        for row in rows_iter:
            values = [str(v).strip() if v is not None else "" for v in row]
            if not any(values):
                continue
            result.append(dict(zip(headers, values, strict=False)))
        wb.close()
        return result
    elif filename.lower().endswith(".csv"):
        import csv
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader]
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 文件")


@router.post("/user/importData")
async def user_import_data(
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:add"))],
    file: UploadFile = File(...),
    update_support: bool = Query(False, alias="updateSupport"),
):
    raw = await file.read()
    filename = file.filename or ""
    rows_data = _parse_import_rows(raw, filename)
    if not rows_data or "userName" not in rows_data[0]:
        raise HTTPException(status_code=400, detail="导入文件必须包含 userName 表头")

    created = 0
    updated = 0
    skipped: list[str] = []
    now = datetime.now()
    for line_no, row in enumerate(rows_data, start=2):
        user_name = (row.get("userName") or "").strip()
        if not user_name:
            skipped.append(f"第 {line_no} 行缺少 userName")
            continue
        user = await db.scalar(select(SysUser).where(SysUser.user_name == user_name))
        if user and not update_support:
            skipped.append(f"{user_name} 已存在")
            continue
        if not user:
            user = SysUser(
                user_name=user_name,
                password=get_password_hash((row.get("password") or "123456").strip() or "123456"),
                create_by=current_name(login_user),
                create_time=now,
                del_flag="0",
            )
            db.add(user)
            created += 1
        else:
            updated += 1
        if row.get("deptId"):
            user.dept_id = int(row["deptId"])
        user.nick_name = (row.get("nickName") or user_name).strip()
        user.email = (row.get("email") or "").strip()
        user.phonenumber = (row.get("phonenumber") or "").strip()
        user.sex = (row.get("sex") or "0").strip()
        user.status = (row.get("status") or "0").strip()
        user.remark = row.get("remark")
        user.update_by = current_name(login_user)
        user.update_time = now
    await db.commit()
    return success()


# UPLOAD_DIR for avatars
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "avatars")
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.get("/user/profile")
async def user_profile(
    login_user: Annotated[LoginUser, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user = login_user.user
    posts = list(
        (await db.execute(
            select(SysPost).join(SysUserPost, SysPost.post_id == SysUserPost.post_id)
            .where(SysUserPost.user_id == user.user_id)
        )).scalars()
    )
    roles = list(
        (await db.execute(
            select(SysRole).join(SysUserRole, SysRole.role_id == SysUserRole.role_id)
            .where(SysUserRole.user_id == user.user_id)
        )).scalars()
    )
    return success(
        data=serialize_user(user),
        roleGroup="、".join(r.role_name for r in roles) or "无角色",
        postGroup="、".join(p.post_name for p in posts) or "无岗位",
        roles=[serialize_role(r) for r in roles],
        postIds=[p.post_id for p in posts],
        roleIds=[r.role_id for r in roles],
    )


@router.put("/user/profile")
async def user_profile_update(
    body: dict[str, Any],
    login_user: Annotated[LoginUser, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user = login_user.user
    for attr, key in [
        ("nick_name", "nickName"),
        ("phonenumber", "phonenumber"),
        ("email", "email"),
        ("sex", "sex"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(user, attr, value)
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    await db.commit()
    return success()


@router.put("/user/profile/updatePwd")
async def user_profile_update_pwd(
    body: dict[str, Any],
    login_user: Annotated[LoginUser, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    old_password = str(field(body, "oldPassword", ""))
    new_password = str(field(body, "newPassword", ""))
    if not verify_password(old_password, login_user.user.password):
        raise HTTPException(status_code=400, detail="旧密码不正确")
    login_user.user.password = get_password_hash(new_password)
    login_user.user.update_by = current_name(login_user)
    login_user.user.update_time = datetime.now()
    await db.commit()
    return success()


@router.post("/user/profile/avatar")
async def user_profile_avatar(
    avatarfile: UploadFile = File(...),
    login_user: LoginUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ext = os.path.splitext(avatarfile.filename or "avatar.png")[1] or ".png"
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    content = await avatarfile.read()
    with open(filepath, "wb") as f:
        f.write(content)
    login_user.user.avatar = f"/uploads/avatars/{filename}"
    login_user.user.update_by = current_name(login_user)
    login_user.user.update_time = datetime.now()
    await db.commit()
    return success(imgUrl=f"/uploads/avatars/{filename}")


@router.get("/user/authRole/{user_id}")
async def user_auth_role_detail(
    user_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    user = await db.get(SysUser, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="数据不存在")
    assigned_ids = set(
        (await db.execute(select(SysUserRole.role_id).where(SysUserRole.user_id == user_id))).scalars()
    )
    all_roles = list((await db.execute(select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort))).scalars())
    return success(
        user=serialize_user(user),
        roles=[{**serialize_role(r), "flag": r.role_id in assigned_ids} for r in all_roles],
    )


@router.get("/user/export")
async def user_export(
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
    user_name: str | None = Query(None, alias="userName"),
    phonenumber: str | None = None,
    status: str | None = None,
    dept_id: int | None = Query(None, alias="deptId"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysUser).options(selectinload(SysUser.dept)).where(SysUser.del_flag == "0").order_by(SysUser.user_id)
    stmt = await apply_data_scope(db, stmt, login_user, SysUser.dept_id, SysUser.user_id)
    if user_name:
        stmt = stmt.where(SysUser.user_name.ilike(f"%{user_name}%"))
    if phonenumber:
        stmt = stmt.where(SysUser.phonenumber.ilike(f"%{phonenumber}%"))
    if status:
        stmt = stmt.where(SysUser.status == status)
    if dept_id:
        child_dept_ids = await collect_child_dept_ids(db, dept_id)
        stmt = stmt.where(SysUser.dept_id.in_([dept_id, *child_dept_ids]))
    stmt = apply_time_range(stmt, SysUser.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars().unique())
    data = [
        [
            str(u.user_id), u.user_name, u.nick_name,
            u.dept.dept_name if u.dept else "",
            u.phonenumber, u.email,
            "男" if u.sex == "0" else "女" if u.sex == "1" else "未知",
            "正常" if u.status == "0" else "停用",
            str(u.create_time or ""), u.remark or ""
        ]
        for u in rows
    ]
    return await xlsx_export(
        "user_export.xlsx",
        ["用户ID", "账号", "昵称", "部门", "手机", "邮箱", "性别", "状态", "创建时间", "备注"],
        data,
    )


@router.get("/user/{user_id}")
async def user_detail(
    user_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:query"))],
):
    user = (
        await db.execute(
            select(SysUser)
            .options(selectinload(SysUser.roles), selectinload(SysUser.posts))
            .where(SysUser.user_id == user_id)
        )
    ).scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="数据不存在")
    role_names = "、".join(role.role_name for role in user.roles) or "无角色"
    post_names = "、".join(post.post_name for post in user.posts) or "无岗位"
    role_ids = [role.role_id for role in user.roles]
    post_ids = [post.post_id for post in user.posts]
    roles = list((await db.execute(select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort))).scalars())
    posts = list((await db.execute(select(SysPost).order_by(SysPost.post_sort))).scalars())
    data = serialize_user(user)
    data["roleNames"] = role_names
    data["postNames"] = post_names
    return success(
        data=data,
        roleIds=role_ids,
        postIds=post_ids,
        roles=[serialize_role(role) for role in roles],
        posts=[serialize_post(post) for post in posts],
    )


@router.post("/user")
async def user_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:add"))],
):
    now = datetime.now()
    password = str(field(body, "password", "123456"))
    await ensure_unique(db, SysUser, SysUser.user_name, field(body, "userName"), "用户账号已存在")
    await ensure_unique(db, SysUser, SysUser.phonenumber, field(body, "phonenumber"), "手机号已存在")
    await ensure_unique(db, SysUser, SysUser.email, field(body, "email"), "邮箱已存在")
    user = SysUser(
        dept_id=field(body, "deptId"),
        user_name=field(body, "userName"),
        nick_name=field(body, "nickName", field(body, "userName")),
        email=field(body, "email", ""),
        phonenumber=field(body, "phonenumber", ""),
        sex=field(body, "sex", "0"),
        password=get_password_hash(password),
        status=field(body, "status", "0"),
        remark=field(body, "remark"),
        create_by=current_name(login_user),
        create_time=now,
    )
    db.add(user)
    await db.flush()
    if has_field(body, "roleIds") or has_field(body, "postIds"):
        await sync_user_roles_posts(
            db,
            user.user_id,
            field(body, "roleIds", []),
            field(body, "postIds", []),
        )
    await db.commit()
    return success()


@router.put("/user")
async def user_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:edit"))],
):
    user_id = int(field(body, "userId"))
    user = await ensure_exists(db, SysUser, user_id)
    if user.is_admin and not login_user.user.is_admin:
        raise HTTPException(status_code=403, detail="不能修改超级管理员")
    await ensure_unique(db, SysUser, SysUser.user_name, field(body, "userName"), "用户账号已存在", SysUser.user_id, user_id)
    await ensure_unique(db, SysUser, SysUser.phonenumber, field(body, "phonenumber"), "手机号已存在", SysUser.user_id, user_id)
    await ensure_unique(db, SysUser, SysUser.email, field(body, "email"), "邮箱已存在", SysUser.user_id, user_id)
    for attr, key in [
        ("dept_id", "deptId"),
        ("user_name", "userName"),
        ("nick_name", "nickName"),
        ("email", "email"),
        ("phonenumber", "phonenumber"),
        ("sex", "sex"),
        ("status", "status"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(user, attr, value)
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    if has_field(body, "roleIds") or has_field(body, "postIds"):
        await sync_user_roles_posts(
            db,
            user.user_id,
            field(body, "roleIds", []),
            field(body, "postIds", []),
        )
    await db.commit()
    return success()


@router.put("/user/resetPwd")
async def user_reset_pwd(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:resetPwd"))],
):
    user = await ensure_exists(db, SysUser, int(field(body, "userId")))
    user.password = get_password_hash(str(field(body, "password")))
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    await db.commit()
    return success()


@router.put("/user/changeStatus")
async def user_change_status(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:user:edit"))],
):
    user = await ensure_exists(db, SysUser, int(field(body, "userId")))
    if user.is_admin:
        raise HTTPException(status_code=400, detail="不能停用超级管理员")
    user.status = str(field(body, "status", user.status))
    user.update_by = current_name(login_user)
    user.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/user/{user_ids}")
async def user_remove(
    user_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:user:remove"))],
):
    ids = [int(item) for item in user_ids.split(",") if item]
    if 1 in ids:
        raise HTTPException(status_code=400, detail="不能删除超级管理员")
    rows = (await db.execute(select(SysUser).where(SysUser.user_id.in_(ids)))).scalars()
    for user in rows:
        user.del_flag = "2"
    await db.commit()
    return success()


@router.put("/user/authRole")
async def user_auth_role(
    user_id: int | None = Query(None, alias="userId"),
    user_id_snake: int | None = Query(None, alias="user_id"),
    role_ids: str = Query("", alias="roleIds"),
    db: AsyncSession = Depends(get_db),
    _login_user: LoginUser = Depends(require_perm("system:user:edit")),
):
    user_id = user_id or user_id_snake
    if user_id is None:
        raise HTTPException(status_code=400, detail="缺少用户ID")
    ids = [int(item) for item in role_ids.split(",") if item]
    await replace_rows(
        db,
        SysUserRole,
        SysUserRole.user_id == user_id,
        [SysUserRole(user_id=user_id, role_id=role_id) for role_id in ids],
    )
    await db.commit()
    return success()


@router.get("/role/list")
async def role_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:role:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    role_name: str | None = Query(None, alias="roleName"),
    role_key: str | None = Query(None, alias="roleKey"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort)
    if role_name:
        stmt = stmt.where(SysRole.role_name.ilike(f"%{role_name}%"))
    if role_key:
        stmt = stmt.where(SysRole.role_key.ilike(f"%{role_key}%"))
    if status:
        stmt = stmt.where(SysRole.status == status)
    stmt = apply_time_range(stmt, SysRole.create_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_role(row) for row in rows], total)


@router.post("/role")
async def role_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:role:add"))],
):
    await ensure_unique(db, SysRole, SysRole.role_name, field(body, "roleName"), "角色名称已存在")
    await ensure_unique(db, SysRole, SysRole.role_key, field(body, "roleKey"), "权限字符已存在")
    role = SysRole(
        role_name=field(body, "roleName"),
        role_key=field(body, "roleKey"),
        role_sort=int(field(body, "roleSort", 0)),
        data_scope=field(body, "dataScope", "1"),
        menu_check_strictly=bool(field(body, "menuCheckStrictly", True)),
        dept_check_strictly=bool(field(body, "deptCheckStrictly", True)),
        status=field(body, "status", "0"),
        remark=field(body, "remark"),
        create_by=current_name(login_user),
        create_time=datetime.now(),
    )
    db.add(role)
    await db.flush()
    if has_field(body, "menuIds"):
        await sync_role_menus(db, role.role_id, field(body, "menuIds", []))
    await db.commit()
    return success()


@router.put("/role")
async def role_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:role:edit"))],
):
    role = await ensure_exists(db, SysRole, int(field(body, "roleId")))
    await ensure_unique(db, SysRole, SysRole.role_name, field(body, "roleName"), "角色名称已存在", SysRole.role_id, role.role_id)
    await ensure_unique(db, SysRole, SysRole.role_key, field(body, "roleKey"), "权限字符已存在", SysRole.role_id, role.role_id)
    if role.is_admin:
        raise HTTPException(status_code=400, detail="不能修改超级管理员角色")
    for attr, key in [
        ("role_name", "roleName"),
        ("role_key", "roleKey"),
        ("role_sort", "roleSort"),
        ("data_scope", "dataScope"),
        ("menu_check_strictly", "menuCheckStrictly"),
        ("dept_check_strictly", "deptCheckStrictly"),
        ("status", "status"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(role, attr, value)
    role.update_by = current_name(login_user)
    role.update_time = datetime.now()
    if has_field(body, "menuIds"):
        await sync_role_menus(db, role.role_id, field(body, "menuIds", []))
    await db.commit()
    await invalidate_tokens_for_role(db, role.role_id)
    return success()


@router.put("/role/dataScope")
async def role_data_scope(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:role:edit"))],
):
    role = await ensure_exists(db, SysRole, int(field(body, "roleId")))
    if role.is_admin:
        raise HTTPException(status_code=400, detail="不能修改超级管理员角色")
    role.data_scope = str(field(body, "dataScope", role.data_scope))
    role.update_by = current_name(login_user)
    role.update_time = datetime.now()
    await sync_role_depts(db, role.role_id, field(body, "deptIds", []))
    await db.commit()
    await invalidate_tokens_for_role(db, role.role_id)
    return success()


@router.put("/role/changeStatus")
async def role_change_status(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:role:edit"))],
):
    role = await ensure_exists(db, SysRole, int(field(body, "roleId")))
    if role.is_admin:
        raise HTTPException(status_code=400, detail="不能停用超级管理员角色")
    role.status = str(field(body, "status", role.status))
    role.update_by = current_name(login_user)
    role.update_time = datetime.now()
    await db.commit()
    await invalidate_tokens_for_role(db, role.role_id)
    return success()


@router.delete("/role/{role_ids}")
async def role_remove(
    role_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:role:remove"))],
):
    ids = [int(item) for item in role_ids.split(",") if item]
    if 1 in ids:
        raise HTTPException(status_code=400, detail="不能删除超级管理员角色")
    for role_id in ids:
        await invalidate_tokens_for_role(db, role_id)
    rows = (await db.execute(select(SysRole).where(SysRole.role_id.in_(ids)))).scalars()
    for role in rows:
        role.del_flag = "2"
    await db.commit()
    return success()


@router.get("/role/optionselect")
async def role_optionselect(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:role:query"))],
):
    rows = list((await db.execute(select(SysRole).where(SysRole.del_flag == "0"))).scalars())
    return success([serialize_role(row) for row in rows])


@router.get("/role/deptTree/{role_id}")
async def role_dept_tree(
    role_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:role:query"))],
):
    depts = list(
        (await db.execute(select(SysDept).where(SysDept.del_flag == "0").order_by(SysDept.order_num))).scalars()
    )
    checked = list(
        (await db.execute(select(SysRoleDept.dept_id).where(SysRoleDept.role_id == role_id))).scalars()
    )
    return success(depts=[dept_tree(node) for node in build_dept_tree(depts, 0)], checkedKeys=checked)


@router.get("/role/authUser/allocatedList")
async def role_allocated_users(
    role_id: int = Query(..., alias="roleId"),
    db: AsyncSession = Depends(get_db),
    pages: tuple[int, int] = Depends(page_params),
    user_name: str | None = Query(None, alias="userName"),
    phonenumber: str | None = None,
    _login_user: LoginUser = Depends(require_perm("system:role:query")),
):
    stmt = (
        select(SysUser)
        .join(SysUserRole, SysUser.user_id == SysUserRole.user_id)
        .where(SysUser.del_flag == "0", SysUserRole.role_id == role_id)
        .order_by(SysUser.user_id)
    )
    if user_name:
        stmt = stmt.where(SysUser.user_name.ilike(f"%{user_name}%"))
    if phonenumber:
        stmt = stmt.where(SysUser.phonenumber.ilike(f"%{phonenumber}%"))
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_user(row) for row in rows], total)


@router.get("/role/authUser/unallocatedList")
async def role_unallocated_users(
    role_id: int = Query(..., alias="roleId"),
    db: AsyncSession = Depends(get_db),
    pages: tuple[int, int] = Depends(page_params),
    user_name: str | None = Query(None, alias="userName"),
    phonenumber: str | None = None,
    _login_user: LoginUser = Depends(require_perm("system:role:query")),
):
    subquery = select(SysUserRole.user_id).where(SysUserRole.role_id == role_id)
    stmt = (
        select(SysUser)
        .where(SysUser.del_flag == "0", SysUser.user_id.not_in(subquery))
        .order_by(SysUser.user_id)
    )
    if user_name:
        stmt = stmt.where(SysUser.user_name.ilike(f"%{user_name}%"))
    if phonenumber:
        stmt = stmt.where(SysUser.phonenumber.ilike(f"%{phonenumber}%"))
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_user(row) for row in rows], total)


@router.put("/role/authUser/selectAll")
async def role_auth_user_select_all(
    role_id: int = Query(..., alias="roleId"),
    user_ids: str = Query("", alias="userIds"),
    db: AsyncSession = Depends(get_db),
    _login_user: LoginUser = Depends(require_perm("system:role:edit")),
):
    ids = [int(item) for item in user_ids.split(",") if item]
    existing = set(
        (
            await db.execute(
                select(SysUserRole.user_id).where(
                    SysUserRole.role_id == role_id,
                    SysUserRole.user_id.in_(ids),
                )
            )
        ).scalars()
    )
    db.add_all(
        [SysUserRole(user_id=user_id, role_id=role_id) for user_id in ids if user_id not in existing]
    )
    await db.commit()
    return success()


@router.delete("/role/authUser/cancelAll")
async def role_auth_user_cancel_all(
    role_id: int = Query(..., alias="roleId"),
    user_ids: str = Query("", alias="userIds"),
    db: AsyncSession = Depends(get_db),
    _login_user: LoginUser = Depends(require_perm("system:role:edit")),
):
    ids = [int(item) for item in user_ids.split(",") if item]
    await db.execute(
        delete(SysUserRole).where(
            SysUserRole.role_id == role_id,
            SysUserRole.user_id.in_(ids),
        )
    )
    await db.commit()
    return success()


@router.get("/role/{role_id}")
async def role_detail(
    role_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:role:query"))],
):
    role = await ensure_exists(db, SysRole, role_id)
    menu_ids = list(
        (await db.execute(select(SysRoleMenu.menu_id).where(SysRoleMenu.role_id == role_id))).scalars()
    )
    dept_ids = list(
        (await db.execute(select(SysRoleDept.dept_id).where(SysRoleDept.role_id == role_id))).scalars()
    )
    return success(data=serialize_role(role), menuIds=menu_ids, deptIds=dept_ids)


@router.get("/menu/list")
async def menu_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:menu:list"))],
    menu_name: str | None = Query(None, alias="menuName"),
    status: str | None = None,
):
    stmt = select(SysMenu).order_by(SysMenu.parent_id, SysMenu.order_num)
    if menu_name:
        stmt = stmt.where(SysMenu.menu_name.ilike(f"%{menu_name}%"))
    if status:
        stmt = stmt.where(SysMenu.status == status)
    rows = list((await db.execute(stmt)).scalars().unique())
    return success([serialize_menu(row) for row in rows])


@router.get("/menu/treeselect")
async def menu_tree_select(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:menu:list"))],
):
    rows = list((await db.execute(select(SysMenu).order_by(SysMenu.parent_id, SysMenu.order_num))).scalars())
    return success([tree_select(node) for node in build_tree(rows, 0)])


@router.get("/menu/roleMenuTreeselect/{role_id}")
async def menu_role_treeselect(
    role_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:role:query"))],
):
    rows = list((await db.execute(select(SysMenu).order_by(SysMenu.parent_id, SysMenu.order_num))).scalars())
    role_menus = list(
        (await db.execute(select(SysRoleMenu.menu_id).where(SysRoleMenu.role_id == role_id))).scalars()
    )
    tree = [tree_select(node) for node in build_tree(rows, 0)]
    return success(checkedKeys=role_menus, menus=tree)


@router.get("/menu/{menu_id}")
async def menu_detail(
    menu_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:menu:query"))],
):
    menu = await ensure_exists(db, SysMenu, menu_id)
    return success(data=serialize_menu(menu))


@router.post("/menu")
async def menu_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:menu:add"))],
):
    same_name = await db.scalar(
        select(func.count()).select_from(SysMenu).where(
            SysMenu.parent_id == int(field(body, "parentId", 0)),
            SysMenu.menu_name == field(body, "menuName"),
        )
    )
    if same_name:
        raise HTTPException(status_code=400, detail="同级菜单名称已存在")
    menu = SysMenu(
        menu_name=field(body, "menuName"),
        parent_id=int(field(body, "parentId", 0)),
        order_num=int(field(body, "orderNum", 0)),
        path=field(body, "path", ""),
        component=field(body, "component"),
        query=field(body, "query"),
        route_name=field(body, "routeName", ""),
        is_frame=int(field(body, "isFrame", 1)),
        is_cache=int(field(body, "isCache", 0)),
        menu_type=field(body, "menuType", ""),
        visible=field(body, "visible", "0"),
        status=field(body, "status", "0"),
        perms=field(body, "perms"),
        icon=field(body, "icon", "#"),
        remark=field(body, "remark", ""),
        create_by=current_name(login_user),
        create_time=datetime.now(),
    )
    db.add(menu)
    await db.commit()
    return success()


@router.put("/menu")
async def menu_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:menu:edit"))],
):
    menu = await ensure_exists(db, SysMenu, int(field(body, "menuId")))
    parent_id_value = int(field(body, "parentId", menu.parent_id))
    same_name = await db.scalar(
        select(func.count()).select_from(SysMenu).where(
            SysMenu.parent_id == parent_id_value,
            SysMenu.menu_name == field(body, "menuName"),
            SysMenu.menu_id != menu.menu_id,
        )
    )
    if same_name:
        raise HTTPException(status_code=400, detail="同级菜单名称已存在")
    for attr, key in [
        ("menu_name", "menuName"),
        ("parent_id", "parentId"),
        ("order_num", "orderNum"),
        ("path", "path"),
        ("component", "component"),
        ("query", "query"),
        ("route_name", "routeName"),
        ("is_frame", "isFrame"),
        ("is_cache", "isCache"),
        ("menu_type", "menuType"),
        ("visible", "visible"),
        ("status", "status"),
        ("perms", "perms"),
        ("icon", "icon"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(menu, attr, value)
    menu.update_by = current_name(login_user)
    menu.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/menu/{menu_ids}")
async def menu_remove(
    menu_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:menu:remove"))],
):
    ids = [int(item) for item in menu_ids.split(",") if item]
    for menu_id in ids:
        children = await db.scalar(select(func.count()).select_from(SysMenu).where(SysMenu.parent_id == menu_id))
        if children:
            raise HTTPException(status_code=400, detail="存在子菜单，不能删除")
    await db.execute(delete(SysMenu).where(SysMenu.menu_id.in_(ids)))
    await db.commit()
    return success()


@router.put("/menu/updateSort")
async def menu_update_sort(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:menu:edit"))],
):
    menu_ids = field(body, "menuIds", "")
    order_nums = field(body, "orderNums", "")
    if not menu_ids or not order_nums:
        raise HTTPException(status_code=400, detail="参数不能为空")
    id_list = [int(x) for x in str(menu_ids).split(",")]
    num_list = [int(x) for x in str(order_nums).split(",")]
    for menu_id, order_num in zip(id_list, num_list):
        menu = await db.get(SysMenu, menu_id)
        if menu:
            menu.order_num = order_num
            menu.update_by = current_name(login_user)
            menu.update_time = datetime.now()
    await db.commit()
    return success()


@router.get("/dept/list")
async def dept_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dept:list"))],
    dept_name: str | None = Query(None, alias="deptName"),
    status: str | None = None,
):
    stmt = select(SysDept).where(SysDept.del_flag == "0").order_by(SysDept.parent_id, SysDept.order_num)
    stmt = await apply_data_scope(db, stmt, login_user, SysDept.dept_id)
    if dept_name:
        stmt = stmt.where(SysDept.dept_name.ilike(f"%{dept_name}%"))
    if status:
        stmt = stmt.where(SysDept.status == status)
    rows = list((await db.execute(stmt)).scalars())
    return success([serialize_dept(row) for row in rows])


@router.get("/dept/treeselect")
async def dept_tree_select(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dept:query"))],
):
    rows = list((await db.execute(select(SysDept).where(SysDept.del_flag == "0").order_by(SysDept.parent_id, SysDept.order_num))).scalars())
    tree = build_dept_tree(rows, 0)
    return success([dept_tree(node) for node in tree])


@router.get("/dept/list/exclude/{dept_id}")
async def dept_list_exclude_child(
    dept_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dept:list"))],
):
    excluded = {dept_id, *(await collect_child_dept_ids(db, dept_id))}
    rows = list(
        (
            await db.execute(
                select(SysDept)
                .where(SysDept.del_flag == "0", SysDept.dept_id.not_in(excluded))
                .order_by(SysDept.parent_id, SysDept.order_num)
            )
        ).scalars()
    )
    return success([serialize_dept(row) for row in rows])


@router.get("/dept/{dept_id}")
async def dept_detail(
    dept_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dept:query"))],
):
    dept = await ensure_exists(db, SysDept, dept_id)
    return success(data=serialize_dept(dept))


@router.post("/dept")
async def dept_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dept:add"))],
):
    parent_id = int(field(body, "parentId", 0))
    dept = SysDept(
        parent_id=parent_id,
        ancestors=await next_ancestors(db, parent_id),
        dept_name=field(body, "deptName"),
        order_num=int(field(body, "orderNum", 0)),
        leader=field(body, "leader"),
        phone=field(body, "phone"),
        email=field(body, "email"),
        status=field(body, "status", "0"),
        del_flag="0",
        create_by=current_name(login_user),
        create_time=datetime.now(),
    )
    db.add(dept)
    await db.commit()
    return success()


@router.put("/dept")
async def dept_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dept:edit"))],
):
    dept = await ensure_exists(db, SysDept, int(field(body, "deptId")))
    parent_id = int(field(body, "parentId", dept.parent_id))
    dept.parent_id = parent_id
    dept.ancestors = await next_ancestors(db, parent_id)
    for attr, key in [
        ("dept_name", "deptName"),
        ("order_num", "orderNum"),
        ("leader", "leader"),
        ("phone", "phone"),
        ("email", "email"),
        ("status", "status"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(dept, attr, value)
    dept.update_by = current_name(login_user)
    dept.update_time = datetime.now()
    await db.commit()
    return success()


@router.put("/dept/updateSort")
async def dept_update_sort(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dept:edit"))],
):
    dept_ids = field(body, "deptIds", "")
    order_nums = field(body, "orderNums", "")
    if not dept_ids or not order_nums:
        raise HTTPException(status_code=400, detail="参数不能为空")
    id_list = [int(x) for x in str(dept_ids).split(",")]
    num_list = [int(x) for x in str(order_nums).split(",")]
    for dept_id, order_num in zip(id_list, num_list):
        dept = await db.get(SysDept, dept_id)
        if dept:
            dept.order_num = order_num
            dept.update_by = current_name(login_user)
            dept.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/dept/{dept_ids}")
async def dept_remove(
    dept_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dept:remove"))],
):
    ids = [int(item) for item in dept_ids.split(",") if item]
    for dept_id in ids:
        children = await db.scalar(
            select(func.count()).select_from(SysDept).where(SysDept.parent_id == dept_id, SysDept.del_flag == "0")
        )
        if children:
            raise HTTPException(status_code=400, detail="存在子部门，不能删除")
    rows = (await db.execute(select(SysDept).where(SysDept.dept_id.in_(ids)))).scalars()
    for dept in rows:
        dept.del_flag = "2"
    await db.commit()
    return success()


@router.get("/post/list")
async def post_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:post:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    post_code: str | None = Query(None, alias="postCode"),
    post_name: str | None = Query(None, alias="postName"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysPost).order_by(SysPost.post_sort)
    if post_code:
        stmt = stmt.where(SysPost.post_code.ilike(f"%{post_code}%"))
    if post_name:
        stmt = stmt.where(SysPost.post_name.ilike(f"%{post_name}%"))
    if status:
        stmt = stmt.where(SysPost.status == status)
    stmt = apply_time_range(stmt, SysPost.create_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_post(row) for row in rows], total)


@router.get("/post/{post_id}")
async def post_detail(
    post_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:post:query"))],
):
    return success(data=serialize_post(await ensure_exists(db, SysPost, post_id)))


@router.post("/post")
async def post_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:post:add"))],
):
    await ensure_unique(db, SysPost, SysPost.post_code, field(body, "postCode"), "岗位编码已存在")
    await ensure_unique(db, SysPost, SysPost.post_name, field(body, "postName"), "岗位名称已存在")
    post = SysPost(
        post_code=field(body, "postCode"),
        post_name=field(body, "postName"),
        post_sort=int(field(body, "postSort", 0)),
        status=field(body, "status", "0"),
        remark=field(body, "remark"),
        create_by=current_name(login_user),
        create_time=datetime.now(),
    )
    db.add(post)
    await db.commit()
    return success()


@router.put("/post")
async def post_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:post:edit"))],
):
    post = await ensure_exists(db, SysPost, int(field(body, "postId")))
    await ensure_unique(db, SysPost, SysPost.post_code, field(body, "postCode"), "岗位编码已存在", SysPost.post_id, post.post_id)
    await ensure_unique(db, SysPost, SysPost.post_name, field(body, "postName"), "岗位名称已存在", SysPost.post_id, post.post_id)
    for attr, key in [
        ("post_code", "postCode"),
        ("post_name", "postName"),
        ("post_sort", "postSort"),
        ("status", "status"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(post, attr, value)
    post.update_by = current_name(login_user)
    post.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/post/{post_ids}")
async def post_remove(
    post_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:post:remove"))],
):
    ids = [int(item) for item in post_ids.split(",") if item]
    await db.execute(delete(SysPost).where(SysPost.post_id.in_(ids)))
    await db.commit()
    return success()


@router.get("/config/list")
async def config_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:config:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    config_name: str | None = Query(None, alias="configName"),
    config_key: str | None = Query(None, alias="configKey"),
    config_type: str | None = Query(None, alias="configType"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysConfig).order_by(SysConfig.config_id)
    if config_name:
        stmt = stmt.where(SysConfig.config_name.ilike(f"%{config_name}%"))
    if config_key:
        stmt = stmt.where(SysConfig.config_key.ilike(f"%{config_key}%"))
    if config_type:
        stmt = stmt.where(SysConfig.config_type == config_type)
    stmt = apply_time_range(stmt, SysConfig.create_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_config(row) for row in rows], total)


@router.get("/config/configKey/{config_key}")
async def config_by_key(
    config_key: str,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    item = await db.scalar(select(SysConfig).where(SysConfig.config_key == config_key))
    return success(item.config_value if item else "")


@router.delete("/config/refreshCache")
async def config_refresh_cache(
    _login_user: Annotated[LoginUser, Depends(require_perm("system:config:remove"))],
):
    return success(msg="参数缓存刷新成功")


@router.get("/config/{config_id}")
async def config_detail(
    config_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:config:query"))],
):
    return success(data=serialize_config(await ensure_exists(db, SysConfig, config_id)))


@router.post("/config")
async def config_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:config:add"))],
):
    await ensure_unique(db, SysConfig, SysConfig.config_key, field(body, "configKey"), "参数键名已存在")
    item = SysConfig(
        config_name=field(body, "configName", ""),
        config_key=field(body, "configKey"),
        config_value=field(body, "configValue", ""),
        config_type=field(body, "configType", "N"),
        remark=field(body, "remark"),
        create_by=current_name(login_user),
        create_time=datetime.now(),
    )
    db.add(item)
    await db.commit()
    return success()


@router.put("/config")
async def config_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:config:edit"))],
):
    item = await ensure_exists(db, SysConfig, int(field(body, "configId")))
    await ensure_unique(db, SysConfig, SysConfig.config_key, field(body, "configKey"), "参数键名已存在", SysConfig.config_id, item.config_id)
    for attr, key in [
        ("config_name", "configName"),
        ("config_key", "configKey"),
        ("config_value", "configValue"),
        ("config_type", "configType"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(item, attr, value)
    item.update_by = current_name(login_user)
    item.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/config/{config_ids}")
async def config_remove(
    config_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:config:remove"))],
):
    ids = [int(item) for item in config_ids.split(",") if item]
    await db.execute(delete(SysConfig).where(SysConfig.config_id.in_(ids)))
    await db.commit()
    return success()


@router.get("/dict/type/list")
async def dict_type_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    dict_name: str | None = Query(None, alias="dictName"),
    dict_type: str | None = Query(None, alias="dictType"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysDictType).order_by(SysDictType.dict_id)
    if dict_name:
        stmt = stmt.where(SysDictType.dict_name.ilike(f"%{dict_name}%"))
    if dict_type:
        stmt = stmt.where(SysDictType.dict_type.ilike(f"%{dict_type}%"))
    if status:
        stmt = stmt.where(SysDictType.status == status)
    stmt = apply_time_range(stmt, SysDictType.create_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_dict_type(row) for row in rows], total)


@router.get("/dict/type/optionselect")
async def dict_type_optionselect(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:list"))],
):
    rows = list((await db.execute(select(SysDictType).order_by(SysDictType.dict_id))).scalars())
    return success(data=[serialize_dict_type(row) for row in rows])


@router.delete("/dict/type/refreshCache")
async def dict_type_refresh_cache(
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:remove"))],
):
    return success(msg="字典缓存刷新成功")


@router.get("/dict/type/{dict_id}")
async def dict_type_detail(
    dict_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:query"))],
):
    return success(data=serialize_dict_type(await ensure_exists(db, SysDictType, dict_id)))


@router.post("/dict/type")
async def dict_type_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dict:add"))],
):
    await ensure_unique(db, SysDictType, SysDictType.dict_type, field(body, "dictType"), "字典类型已存在")
    item = SysDictType(
        dict_name=field(body, "dictName", ""),
        dict_type=field(body, "dictType"),
        status=field(body, "status", "0"),
        remark=field(body, "remark"),
        create_by=current_name(login_user),
        create_time=datetime.now(),
    )
    db.add(item)
    await db.commit()
    return success()


@router.put("/dict/type")
async def dict_type_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dict:edit"))],
):
    item = await ensure_exists(db, SysDictType, int(field(body, "dictId")))
    await ensure_unique(db, SysDictType, SysDictType.dict_type, field(body, "dictType"), "字典类型已存在", SysDictType.dict_id, item.dict_id)
    for attr, key in [
        ("dict_name", "dictName"),
        ("dict_type", "dictType"),
        ("status", "status"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(item, attr, value)
    item.update_by = current_name(login_user)
    item.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/dict/type/{dict_ids}")
async def dict_type_remove(
    dict_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:remove"))],
):
    ids = [int(item) for item in dict_ids.split(",") if item]
    await db.execute(delete(SysDictType).where(SysDictType.dict_id.in_(ids)))
    await db.commit()
    return success()


@router.get("/dict/data/list")
async def dict_data_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    dict_type: str | None = Query(None, alias="dictType"),
    dict_label: str | None = Query(None, alias="dictLabel"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysDictData).order_by(SysDictData.dict_sort)
    if dict_type:
        stmt = stmt.where(SysDictData.dict_type == dict_type)
    if dict_label:
        stmt = stmt.where(SysDictData.dict_label.ilike(f"%{dict_label}%"))
    if status:
        stmt = stmt.where(SysDictData.status == status)
    stmt = apply_time_range(stmt, SysDictData.create_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([serialize_dict_data(row) for row in rows], total)


@router.get("/dict/data/type/{dict_type}")
async def dict_data_by_type(dict_type: str, db: Annotated[AsyncSession, Depends(get_db)]):
    rows = list(
        (
            await db.execute(
                select(SysDictData)
                .where(SysDictData.dict_type == dict_type, SysDictData.status == "0")
                .order_by(SysDictData.dict_sort)
            )
        ).scalars()
    )
    return success([serialize_dict_data(row) for row in rows])


@router.get("/dict/data/{dict_code}")
async def dict_data_detail(
    dict_code: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:query"))],
):
    return success(data=serialize_dict_data(await ensure_exists(db, SysDictData, dict_code)))


@router.post("/dict/data")
async def dict_data_add(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dict:add"))],
):
    item = SysDictData(
        dict_sort=int(field(body, "dictSort", 0)),
        dict_label=field(body, "dictLabel", ""),
        dict_value=field(body, "dictValue", ""),
        dict_type=field(body, "dictType", ""),
        css_class=field(body, "cssClass"),
        list_class=field(body, "listClass"),
        is_default=field(body, "isDefault", "N"),
        status=field(body, "status", "0"),
        remark=field(body, "remark"),
        create_by=current_name(login_user),
        create_time=datetime.now(),
    )
    db.add(item)
    await db.commit()
    return success()


@router.put("/dict/data")
async def dict_data_edit(
    body: dict[str, Any],
    db: Annotated[AsyncSession, Depends(get_db)],
    login_user: Annotated[LoginUser, Depends(require_perm("system:dict:edit"))],
):
    item = await ensure_exists(db, SysDictData, int(field(body, "dictCode")))
    for attr, key in [
        ("dict_sort", "dictSort"),
        ("dict_label", "dictLabel"),
        ("dict_value", "dictValue"),
        ("dict_type", "dictType"),
        ("css_class", "cssClass"),
        ("list_class", "listClass"),
        ("is_default", "isDefault"),
        ("status", "status"),
        ("remark", "remark"),
    ]:
        value = field(body, key)
        if value is not None:
            setattr(item, attr, value)
    item.update_by = current_name(login_user)
    item.update_time = datetime.now()
    await db.commit()
    return success()


@router.delete("/dict/data/{dict_codes}")
async def dict_data_remove(
    dict_codes: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:dict:remove"))],
):
    ids = [int(item) for item in dict_codes.split(",") if item]
    await db.execute(delete(SysDictData).where(SysDictData.dict_code.in_(ids)))
    await db.commit()
    return success()


monitor_router = APIRouter(prefix="/monitor", tags=["monitor"])


_operlog_sort_map = {
    "operName": SysOperLog.oper_name,
    "operTime": SysOperLog.oper_time,
    "costTime": SysOperLog.cost_time,
}


@monitor_router.get("/operlog/list")
async def operlog_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:operlog:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    title: str | None = None,
    oper_name: str | None = Query(None, alias="operName"),
    oper_ip: str | None = Query(None, alias="operIp"),
    business_type: int | None = Query(None, alias="businessType"),
    status: int | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
    order_by_column: str | None = Query(None, alias="orderByColumn"),
    is_asc: str | None = Query(None, alias="isAsc"),
):
    sort_col = _operlog_sort_map.get(order_by_column) if order_by_column else None
    if sort_col is not None and is_asc == "asc":
        stmt = select(SysOperLog).order_by(sort_col.asc())
    elif sort_col is not None:
        stmt = select(SysOperLog).order_by(sort_col.desc())
    else:
        stmt = select(SysOperLog).order_by(SysOperLog.oper_id.desc())
    if title:
        stmt = stmt.where(SysOperLog.title.ilike(f"%{title}%"))
    if oper_name:
        stmt = stmt.where(SysOperLog.oper_name.ilike(f"%{oper_name}%"))
    if oper_ip:
        stmt = stmt.where(SysOperLog.oper_ip.ilike(f"%{oper_ip}%"))
    if business_type is not None:
        stmt = stmt.where(SysOperLog.business_type == business_type)
    if status is not None:
        stmt = stmt.where(SysOperLog.status == status)
    stmt = apply_time_range(stmt, SysOperLog.oper_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([model_dict(row) for row in rows], total)


@monitor_router.delete("/operlog/clean")
async def operlog_clean(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:operlog:remove"))],
):
    await db.execute(delete(SysOperLog))
    await db.commit()
    return success()


@monitor_router.delete("/operlog/{oper_ids}")
async def operlog_remove(
    oper_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:operlog:remove"))],
):
    ids = [int(item) for item in oper_ids.split(",") if item]
    await db.execute(delete(SysOperLog).where(SysOperLog.oper_id.in_(ids)))
    await db.commit()
    return success()


_logininfor_sort_map = {
    "userName": SysLogininfor.user_name,
    "loginTime": SysLogininfor.login_time,
}


@monitor_router.get("/logininfor/list")
async def logininfor_list(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:logininfor:list"))],
    pages: Annotated[tuple[int, int], Depends(page_params)],
    user_name: str | None = Query(None, alias="userName"),
    ipaddr: str | None = None,
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
    order_by_column: str | None = Query(None, alias="orderByColumn"),
    is_asc: str | None = Query(None, alias="isAsc"),
):
    sort_col = _logininfor_sort_map.get(order_by_column) if order_by_column else None
    if sort_col is not None and is_asc == "asc":
        stmt = select(SysLogininfor).order_by(sort_col.asc())
    elif sort_col is not None:
        stmt = select(SysLogininfor).order_by(sort_col.desc())
    else:
        stmt = select(SysLogininfor).order_by(SysLogininfor.info_id.desc())
    if user_name:
        stmt = stmt.where(SysLogininfor.user_name.ilike(f"%{user_name}%"))
    if ipaddr:
        stmt = stmt.where(SysLogininfor.ipaddr.ilike(f"%{ipaddr}%"))
    if status:
        stmt = stmt.where(SysLogininfor.status == status)
    stmt = apply_time_range(stmt, SysLogininfor.login_time, begin_time, end_time)
    rows, total = await page_query(db, stmt, *pages)
    return table([model_dict(row) for row in rows], total)


@monitor_router.delete("/logininfor/clean")
async def logininfor_clean(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:logininfor:remove"))],
):
    await db.execute(delete(SysLogininfor))
    await db.commit()
    return success()


@monitor_router.delete("/logininfor/{info_ids}")
async def logininfor_remove(
    info_ids: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:logininfor:remove"))],
):
    ids = [int(item) for item in info_ids.split(",") if item]
    await db.execute(delete(SysLogininfor).where(SysLogininfor.info_id.in_(ids)))
    await db.commit()
    return success()


@monitor_router.get("/logininfor/unlock/{user_name}")
async def logininfor_unlock(
    user_name: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:logininfor:unlock"))],
):
    user = await db.scalar(select(SysUser).where(SysUser.user_name == user_name))
    if user is None:
        raise HTTPException(status_code=404, detail=f"用户 {user_name} 不存在")
    return success(msg=f"用户 {user_name} 解锁成功")


async def xlsx_export(filename: str, headers: list[str], rows: list[list[str]]) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_data in rows:
            cell_value = str(row_data[col_idx - 1]) if col_idx - 1 < len(row_data) else ""
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/role/export")
async def role_export(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:role:query"))],
    role_name: str | None = Query(None, alias="roleName"),
    role_key: str | None = Query(None, alias="roleKey"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysRole).where(SysRole.del_flag == "0").order_by(SysRole.role_sort)
    if role_name:
        stmt = stmt.where(SysRole.role_name.ilike(f"%{role_name}%"))
    if role_key:
        stmt = stmt.where(SysRole.role_key.ilike(f"%{role_key}%"))
    if status:
        stmt = stmt.where(SysRole.status == status)
    stmt = apply_time_range(stmt, SysRole.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [
            str(r.role_id), r.role_name, r.role_key, str(r.role_sort),
            data_scope_label_str(r.data_scope),
            "正常" if r.status == "0" else "停用",
            str(r.create_time or ""), r.remark or ""
        ]
        for r in rows
    ]
    return await xlsx_export(
        "role_export.xlsx",
        ["角色ID", "角色名称", "权限字符", "排序", "数据范围", "状态", "创建时间", "备注"],
        data,
    )


@router.get("/post/export")
async def post_export(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:post:query"))],
    post_code: str | None = Query(None, alias="postCode"),
    post_name: str | None = Query(None, alias="postName"),
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysPost).order_by(SysPost.post_sort)
    if post_code:
        stmt = stmt.where(SysPost.post_code.ilike(f"%{post_code}%"))
    if post_name:
        stmt = stmt.where(SysPost.post_name.ilike(f"%{post_name}%"))
    if status:
        stmt = stmt.where(SysPost.status == status)
    stmt = apply_time_range(stmt, SysPost.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(p.post_id), p.post_code, p.post_name, str(p.post_sort),
         "正常" if p.status == "0" else "停用",
         str(p.create_time or ""), p.remark or ""]
        for p in rows
    ]
    return await xlsx_export(
        "post_export.xlsx",
        ["岗位ID", "岗位编码", "岗位名称", "排序", "状态", "创建时间", "备注"],
        data,
    )


@router.get("/config/export")
async def config_export(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("system:config:query"))],
    config_name: str | None = Query(None, alias="configName"),
    config_key: str | None = Query(None, alias="configKey"),
    config_type: str | None = Query(None, alias="configType"),
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysConfig).order_by(SysConfig.config_id)
    if config_name:
        stmt = stmt.where(SysConfig.config_name.ilike(f"%{config_name}%"))
    if config_key:
        stmt = stmt.where(SysConfig.config_key.ilike(f"%{config_key}%"))
    if config_type:
        stmt = stmt.where(SysConfig.config_type == config_type)
    stmt = apply_time_range(stmt, SysConfig.create_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(c.config_id), c.config_name or "", c.config_key,
         c.config_value or "", "是" if c.config_type == "Y" else "否",
         str(c.create_time or ""), c.remark or ""]
        for c in rows
    ]
    return await xlsx_export(
        "config_export.xlsx",
        ["参数ID", "参数名称", "参数键名", "参数键值", "系统内置", "创建时间", "备注"],
        data,
    )


@monitor_router.get("/operlog/export")
async def operlog_export(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:operlog:list"))],
    title: str | None = None,
    oper_name: str | None = Query(None, alias="operName"),
    business_type: int | None = Query(None, alias="businessType"),
    status: int | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysOperLog).order_by(SysOperLog.oper_id.desc())
    if title:
        stmt = stmt.where(SysOperLog.title.ilike(f"%{title}%"))
    if oper_name:
        stmt = stmt.where(SysOperLog.oper_name.ilike(f"%{oper_name}%"))
    if business_type is not None:
        stmt = stmt.where(SysOperLog.business_type == business_type)
    if status is not None:
        stmt = stmt.where(SysOperLog.status == status)
    stmt = apply_time_range(stmt, SysOperLog.oper_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(o.oper_id), o.title or "", business_type_label_str(o.business_type or 0),
         o.request_method or "", o.oper_name or "", o.oper_url or "",
         o.oper_ip or "", str(o.oper_time or ""), str(o.cost_time or 0),
         "成功" if o.status == 0 else "失败", o.error_msg or ""]
        for o in rows
    ]
    return await xlsx_export(
        "operlog_export.xlsx",
        ["日志编号", "操作模块", "业务类型", "请求方式", "操作人员", "请求URL", "IP地址", "操作时间", "耗时(ms)", "状态", "错误信息"],
        data,
    )


@monitor_router.get("/logininfor/export")
async def logininfor_export(
    db: Annotated[AsyncSession, Depends(get_db)],
    _login_user: Annotated[LoginUser, Depends(require_perm("monitor:logininfor:list"))],
    user_name: str | None = Query(None, alias="userName"),
    ipaddr: str | None = None,
    status: str | None = None,
    begin_time: str | None = Query(None, alias="beginTime"),
    end_time: str | None = Query(None, alias="endTime"),
):
    stmt = select(SysLogininfor).order_by(SysLogininfor.info_id.desc())
    if user_name:
        stmt = stmt.where(SysLogininfor.user_name.ilike(f"%{user_name}%"))
    if ipaddr:
        stmt = stmt.where(SysLogininfor.ipaddr.ilike(f"%{ipaddr}%"))
    if status:
        stmt = stmt.where(SysLogininfor.status == status)
    stmt = apply_time_range(stmt, SysLogininfor.login_time, begin_time, end_time)
    rows = list((await db.execute(stmt)).scalars())
    data = [
        [str(l.info_id), l.user_name or "", l.ipaddr or "", l.login_location or "",
         l.browser or "", l.os or "", "成功" if l.status == "0" else "失败",
         l.msg or "", str(l.login_time or "")]
        for l in rows
    ]
    return await xlsx_export(
        "logininfor_export.xlsx",
        ["访问编号", "用户账号", "IP地址", "登录地点", "浏览器", "操作系统", "状态", "提示消息", "登录时间"],
        data,
    )


async def sync_user_roles_posts(
    db: AsyncSession,
    user_id: int,
    role_ids: list[int] | None,
    post_ids: list[int] | None,
) -> None:
    await replace_rows(
        db,
        SysUserRole,
        SysUserRole.user_id == user_id,
        [SysUserRole(user_id=user_id, role_id=int(role_id)) for role_id in role_ids or []],
    )
    await replace_rows(
        db,
        SysUserPost,
        SysUserPost.user_id == user_id,
        [SysUserPost(user_id=user_id, post_id=int(post_id)) for post_id in post_ids or []],
    )


async def sync_role_menus(db: AsyncSession, role_id: int, menu_ids: list[int] | None) -> None:
    await replace_rows(
        db,
        SysRoleMenu,
        SysRoleMenu.role_id == role_id,
        [SysRoleMenu(role_id=role_id, menu_id=int(menu_id)) for menu_id in menu_ids or []],
    )


async def sync_role_depts(db: AsyncSession, role_id: int, dept_ids: list[int] | None) -> None:
    await replace_rows(
        db,
        SysRoleDept,
        SysRoleDept.role_id == role_id,
        [SysRoleDept(role_id=role_id, dept_id=int(dept_id)) for dept_id in dept_ids or []],
    )


async def next_ancestors(db: AsyncSession, parent_id: int) -> str:
    if parent_id == 0:
        return "0"
    parent = await db.get(SysDept, parent_id)
    if parent is None:
        return "0"
    return f"{parent.ancestors},{parent.dept_id}"


async def collect_child_dept_ids(db: AsyncSession, dept_id: int) -> list[int]:
    rows = list((await db.execute(select(SysDept).where(SysDept.del_flag == "0"))).scalars())
    result: list[int] = []

    def collect(parent_id: int) -> None:
        for dept in rows:
            if dept.parent_id == parent_id:
                result.append(dept.dept_id)
                collect(dept.dept_id)

    collect(dept_id)
    return result


def build_dept_tree(depts: list[SysDept], parent_id: int = 0) -> list[SysDept]:
    nodes = [dept for dept in depts if dept.parent_id == parent_id]
    for node in nodes:
        setattr(node, "children", build_dept_tree(depts, node.dept_id))
    return nodes


def dept_tree(dept: SysDept) -> dict:
    return {
        "id": dept.dept_id,
        "label": dept.dept_name,
        "children": [dept_tree(child) for child in getattr(dept, "children", [])],
    }


def model_dict(obj) -> dict:
    return {col.name: getattr(obj, col.name) for col in obj.__table__.columns}


def serialize_user(user: SysUser) -> dict:
    return {
        "userId": user.user_id,
        "deptId": user.dept_id,
        "deptName": user.dept.dept_name if user.dept else "",
        "dept": {"deptName": user.dept.dept_name} if user.dept else None,
        "userName": user.user_name,
        "nickName": user.nick_name,
        "email": user.email,
        "phonenumber": user.phonenumber,
        "sex": user.sex,
        "avatar": user.avatar,
        "status": user.status,
        "remark": user.remark,
        "createTime": user.create_time,
    }


def serialize_role(role: SysRole) -> dict:
    data = model_dict(role)
    data.update({
        "roleId": role.role_id,
        "roleName": role.role_name,
        "roleKey": role.role_key,
        "dataScope": role.data_scope,
        "roleSort": role.role_sort,
        "createTime": role.create_time,
        "updateTime": role.update_time,
    })
    return data


def serialize_menu(menu: SysMenu) -> dict:
    return {
        "menuId": menu.menu_id,
        "menuName": menu.menu_name,
        "parentId": menu.parent_id,
        "orderNum": menu.order_num,
        "path": menu.path,
        "component": menu.component,
        "query": menu.query,
        "routeName": menu.route_name,
        "isFrame": menu.is_frame,
        "isCache": menu.is_cache,
        "menuType": menu.menu_type,
        "visible": menu.visible,
        "status": menu.status,
        "perms": menu.perms,
        "icon": menu.icon,
        "remark": menu.remark,
        "createTime": menu.create_time,
        "updateTime": menu.update_time,
    }


def serialize_dept(dept: SysDept) -> dict:
    return {
        "deptId": dept.dept_id,
        "parentId": dept.parent_id,
        "ancestors": dept.ancestors,
        "deptName": dept.dept_name,
        "orderNum": dept.order_num,
        "leader": dept.leader,
        "phone": dept.phone,
        "email": dept.email,
        "status": dept.status,
        "createTime": dept.create_time,
        "updateTime": dept.update_time,
    }


def serialize_post(post: SysPost) -> dict:
    data = model_dict(post)
    data.update({
        "postId": post.post_id,
        "postCode": post.post_code,
        "postName": post.post_name,
        "postSort": post.post_sort,
        "createTime": post.create_time,
        "updateTime": post.update_time,
    })
    return data


def serialize_config(config: SysConfig) -> dict:
    data = model_dict(config)
    data.update({
        "configId": config.config_id,
        "configName": config.config_name,
        "configKey": config.config_key,
        "configValue": config.config_value,
        "configType": config.config_type,
        "createTime": config.create_time,
        "updateTime": config.update_time,
    })
    return data


def serialize_dict_type(item: SysDictType) -> dict:
    data = model_dict(item)
    data.update({
        "dictId": item.dict_id,
        "dictName": item.dict_name,
        "dictType": item.dict_type,
        "createTime": item.create_time,
        "updateTime": item.update_time,
    })
    return data


def serialize_dict_data(item: SysDictData) -> dict:
    data = model_dict(item)
    data.update({
        "dictCode": item.dict_code,
        "dictSort": item.dict_sort,
        "dictLabel": item.dict_label,
        "dictValue": item.dict_value,
        "dictType": item.dict_type,
        "cssClass": item.css_class,
        "listClass": item.list_class,
        "isDefault": item.is_default,
        "createTime": item.create_time,
        "updateTime": item.update_time,
    })
    return data


def tree_select(menu: SysMenu) -> dict:
    return {
        "id": menu.menu_id,
        "label": menu.menu_name,
        "children": [tree_select(child) for child in getattr(menu, "children", [])],
    }


def data_scope_label_str(value: str) -> str:
    return {
        "1": "全部数据权限",
        "2": "自定义数据权限",
        "3": "本部门数据权限",
        "4": "本部门及以下数据权限",
        "5": "仅本人数据权限",
    }.get(value, value)


def business_type_label_str(value: int) -> str:
    return {
        0: "其他",
        1: "新增",
        2: "修改",
        3: "删除",
        4: "授权",
        5: "导出",
        6: "导入",
        7: "强退",
        8: "清空",
        9: "生成代码",
    }.get(value, str(value))
